import os
import re

import requests
from openpyxl import load_workbook


BASE_URL = "https://v2.koko.kg"


def normalize_name(value):
    return re.sub(r"\s+", "", (value or "")).strip()


def sanitize_person_name(value):
    cleaned = str(value or "").strip()
    cleaned = re.sub(r'[\\/:*?"<>|]+', "", cleaned)
    cleaned = re.sub(r"\s+", "", cleaned)
    return cleaned


def build_passport_filename(person_name):
    return f"{sanitize_person_name(person_name)}_护照.jpg"


def build_portrait_filename(person_name, source_path=""):
    ext = os.path.splitext(source_path or "")[1].lower()
    if ext not in {".jpg", ".jpeg", ".png"}:
        ext = ".jpg"
    return f"{sanitize_person_name(person_name)}_证件照{ext}"


def parse_excel_rows(excel_path):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        company_code = (row[1] or "").strip() if isinstance(row[1], str) else row[1]
        company_name = (row[2] or "").strip() if isinstance(row[2], str) else row[2]
        person_name = (row[5] or "").strip() if isinstance(row[5], str) else row[5]
        visa_type_text = (row[7] or "").strip() if isinstance(row[7], str) else row[7]

        if not person_name and not visa_type_text and not company_name:
            continue

        rows.append(
            {
                "row_index": idx,
                "company_code": str(company_code or "").strip(),
                "company_name": str(company_name or "").strip(),
                "person_name": str(person_name or "").strip(),
                "visa_type_text": str(visa_type_text or "").strip(),
            }
        )
    return rows


def parse_visa_type_text(value):
    text = (value or "").strip()
    if not text:
        raise ValueError("Excel 里的类型为空")

    is_super_urgent = "超级加急" in text

    if "普通" in text:
        service_level = "standard"
    elif is_super_urgent or "加急" in text:
        service_level = "urgent"
    else:
        raise ValueError(f"无法识别服务等级：{text}")

    if "多次" in text:
        entry_type = "multiple"
    elif "单次" in text:
        entry_type = "single"
    else:
        raise ValueError(f"无法识别入境类型：{text}")

    month_match = re.search(r"([123])\s*个?月", text)
    if not month_match:
        raise ValueError(f"无法识别有效期：{text}")
    months = int(month_match.group(1))

    return {
        "service_level": service_level,
        "is_super_urgent": is_super_urgent,
        "entry_type": entry_type,
        "months": months,
    }


def find_person_folder(root_dir, person_name):
    target = normalize_name(person_name)
    for entry in os.listdir(root_dir):
        path = os.path.join(root_dir, entry)
        if os.path.isdir(path) and normalize_name(entry) == target:
            return path
    return None


def choose_passport_and_photo(folder_path, looks_like_passport_image):
    files = sorted(os.listdir(folder_path))
    images = []
    for name in files:
        path = os.path.join(folder_path, name)
        if not os.path.isfile(path):
            continue
        lower = name.lower()
        if lower.endswith((".jpg", ".jpeg", ".png")):
            images.append(path)

    passport_file = None
    passport_image = None
    photo_file = None

    for path in images:
        lower_name = os.path.basename(path).lower()
        if "_护照" in lower_name:
            passport_file = path
            passport_image = path
            break

    for path in images:
        if passport_file is not None:
            break
        lower_name = os.path.basename(path).lower()
        if "护照" in lower_name or "passport" in lower_name:
            passport_file = path
            passport_image = path
            break

    if passport_file is None:
        for path in images:
            try:
                from PIL import Image

                img = Image.open(path).convert("RGB")
                if looks_like_passport_image(img):
                    passport_file = path
                    passport_image = path
                    break
            except Exception:
                continue

    photo_hints = ("证件", "photo", "微信", "qq")
    for path in images:
        if path == passport_image:
            continue
        lower_name = os.path.basename(path).lower()
        if "_证件照" in lower_name:
            photo_file = path
            break

    for path in images:
        if photo_file is not None:
            break
        if path == passport_image:
            continue
        lower_name = os.path.basename(path).lower()
        if any(hint in lower_name for hint in photo_hints):
            photo_file = path
            break

    if photo_file is None:
        for path in images:
            if path != passport_image:
                photo_file = path
                break

    return passport_file, photo_file


class KokoVisaClient:
    def __init__(self, api_key):
        self.session = requests.Session()
        self.session.headers.update({"X-API-Key": api_key})

    def _request(self, method, path, **kwargs):
        response = self.session.request(method, BASE_URL + path, timeout=60, **kwargs)
        if response.ok:
            try:
                return response.json()
            except Exception:
                return {}

        message = f"HTTP {response.status_code}"
        try:
            payload = response.json()
            if isinstance(payload, dict):
                detail = payload.get("message") or payload.get("error") or payload.get("detail")
                if detail:
                    message = f"{message}: {detail}"
        except Exception:
            if response.text:
                message = f"{message}: {response.text[:200]}"
        raise RuntimeError(message)

    def search_company(self, company_name):
        payload = self._request("GET", "/api/open/companies", params={"search": company_name})
        items = payload.get("data") or []
        exact = [item for item in items if normalize_name(item.get("name", "")) == normalize_name(company_name)]
        return exact[0] if exact else (items[0] if items else None), payload

    def search_company_payload(self, company_name):
        params = {"search": company_name} if company_name else {}
        return self._request("GET", "/api/open/companies", params=params)

    def search_passports(self, search_text):
        return self._request("GET", "/api/open/passports", params={"search": search_text, "page": 1, "per_page": 50})

    def get_passport_detail(self, passport_id):
        return self._request("GET", f"/api/open/passports/{passport_id}")

    def list_visas(self, *, status="", category="", page=1, per_page=50):
        params = {}
        if status:
            params["status"] = status
        if category:
            params["category"] = category
        params["page"] = page
        params["per_page"] = per_page
        return self._request("GET", "/api/open/visa/list", params=params)

    def find_visas_by_passport_number(self, passport_number, *, status="", category="", per_page=100, max_pages=20):
        target = normalize_name(passport_number).upper()
        if not target:
            raise ValueError("护照号不能为空")

        matched = []
        last_payload = {"data": [], "meta": {"page": 1, "per_page": per_page, "total": 0, "total_pages": 0}}

        for page in range(1, max_pages + 1):
            payload = self.list_visas(status=status, category=category, page=page, per_page=per_page)
            last_payload = payload
            items = payload.get("data") or []

            for item in items:
                item_passport = normalize_name(item.get("passport_number", "")).upper()
                if item_passport == target:
                    matched.append(item)

            meta = payload.get("meta") or {}
            total_pages = int(meta.get("total_pages") or 0)
            if total_pages and page >= total_pages:
                break
            if not items:
                break

        merged_meta = dict(last_payload.get("meta") or {})
        merged_meta["matched_total"] = len(matched)
        return {"data": matched, "meta": merged_meta}

    def get_visa_detail(self, application_id):
        return self._request("GET", f"/api/open/visa/{application_id}")

    def list_tags(self):
        return self._request("GET", "/api/open/tags")

    def add_tag_to_visa(self, application_id, tag_id):
        return self._request("POST", f"/api/open/visa/{application_id}/tags/{tag_id}")

    def submit_visa(self, *, passport_path, photo_path, visa_category, company_id, entry_type, months, service_level, is_super_urgent=False):
        data = {
            "visa_category": visa_category,
            "company_id": company_id,
            "entry_type": entry_type,
            "months": str(months),
            "service_level": service_level,
            "is_super_urgent": "true" if is_super_urgent else "false",
        }

        file_handles = [open(passport_path, "rb")]
        files = [("file", (os.path.basename(passport_path), file_handles[0]))]
        try:
            if photo_path:
                file_handles.append(open(photo_path, "rb"))
                files.append(("photo", (os.path.basename(photo_path), file_handles[-1])))
            payload = self._request("POST", "/api/open/visa/submit", data=data, files=files)
            return payload.get("data") or {}, payload
        finally:
            for handle in file_handles:
                handle.close()
